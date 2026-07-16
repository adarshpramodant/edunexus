import sys
import os
import io
import csv
import uuid
import time
import json
import threading
import math
from datetime import datetime, date
from flask import Blueprint, request, jsonify, send_file

# Adjust path to import db and other modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from db import get_db_connection
from auth_middleware import token_required
from activity_logger import log_activity

import openpyxl
import xlrd
from odf import opendocument
from odf.table import Table, TableRow, TableCell
from odf.text import P

import requests

import openpyxl.writer.excel

import mimetypes

import traceback

import shutil

import ipaddress

import urllib.parse

import logging

import glob

import socket

import decimal

import re

# Initialize import blueprint
import_bp = Blueprint('import', __name__, url_prefix='/api/import')

# ─────────────────────────────────────────────────────────────────────────────
# PARSING SERVICES (Pure Python, Memory Efficient)
# ─────────────────────────────────────────────────────────────────────────────

def parse_csv(file_bytes):
    text = file_bytes.decode('utf-8', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return rows

def parse_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = []
    for r in sheet.iter_rows(values_only=True):
        rows.append(list(r))
    return rows

def parse_xls(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheet = wb.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append(sheet.row_values(r))
    return rows

def parse_ods(file_bytes):
    doc = opendocument.load(io.BytesIO(file_bytes))
    tables = doc.getElementsByType(Table)
    if not tables:
        return []
    sheet = tables[0]
    rows = []
    for row in sheet.getElementsByType(TableRow):
        row_cells = []
        for cell in row.getElementsByType(TableCell):
            # Check number of columns repeated
            number_repeated = cell.getAttribute("numbercolumnsrepeated")
            repeat = int(number_repeated) if number_repeated else 1
            
            text_p = cell.getElementsByType(P)
            val = ""
            if text_p:
                val = "".join([str(t) for t in text_p[0].childNodes])
                
            for _ in range(repeat):
                row_cells.append(val)
        # Skip trailing empty cells to optimize ODS parser
        while row_cells and row_cells[-1] == "":
            row_cells.pop()
        if any(c != "" for c in row_cells):
            rows.append(row_cells)
    return rows

def parse_file(file_bytes, file_name):
    ext = file_name.split('.')[-1].lower()
    if ext == 'csv':
        rows = parse_csv(file_bytes)
    elif ext == 'xlsx':
        rows = parse_xlsx(file_bytes)
    elif ext == 'xls':
        rows = parse_xls(file_bytes)
    elif ext == 'ods':
        rows = parse_ods(file_bytes)
    else:
        raise ValueError("Unsupported file format. Please upload CSV, XLSX, XLS, or ODS.")
    
    if not rows or len(rows) < 1:
        return [], []
        
    # Extract headers
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    
    # Extract records
    records = []
    for r_idx, row in enumerate(rows[1:], start=2):
        if not any(x is not None and str(x).strip() != "" for x in row):
            continue
        record = {'_row_number': r_idx}
        for idx, h in enumerate(headers):
            if not h:
                continue
            val = row[idx] if idx < len(row) else None
            record[h] = val
        records.append(record)
        
    return headers, records

# ─────────────────────────────────────────────────────────────────────────────
# AI AUTO-DETECTION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def detect_column_mappings(headers, import_type=None):
    """
    Scans HeaderAliases in the database to map columns automatically.
    Returns (detected_mapping, confidence_score, detected_type)
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT db_field, alias FROM HeaderAliases")
    aliases = cur.fetchall()
    conn.close()
    
    # Create alias maps
    alias_map = {}
    for r in aliases:
        fld = r['db_field']
        al = r['alias'].strip().lower()
        if fld not in alias_map:
            alias_map[fld] = []
        alias_map[fld].append(al)
        
    mapping = {}
    matched_cols = set()
    
    # Normalize headers
    norm_headers = [str(h).strip().lower() for h in headers]
    
    # Attempt to match each db_field
    for fld, alias_list in alias_map.items():
        for alias in alias_list:
            for idx, nh in enumerate(norm_headers):
                if nh == alias or nh.replace('_', ' ') == alias or alias in nh:
                    mapping[fld] = headers[idx]
                    matched_cols.add(headers[idx])
                    break
            if fld in mapping:
                break
                
    # Detect Import Type if not specified
    if not import_type or import_type == 'auto':
        has_marks = any(f in mapping for f in ['cia1', 'cia2', 'cia3', 'assignment', 'lab', 'project', 'quiz', 'seminar', 'total', 'grade'])
        has_attendance = 'attendance_status' in mapping
        
        if has_marks and has_attendance:
            import_type = 'combined'
        elif has_attendance:
            import_type = 'attendance'
        else:
            import_type = 'marks'
            
    # Calculate confidence score
    expected_matches = 2 # register_number + status/mark
    actual_matches = 0
    if 'register_number' in mapping:
        actual_matches += 1
    if import_type == 'attendance' and 'attendance_status' in mapping:
        actual_matches += 1
    elif import_type == 'marks' and any(f in mapping for f in ['cia1', 'cia2', 'cia3', 'assignment', 'lab', 'total']):
        actual_matches += 1
    elif import_type == 'combined' and 'attendance_status' in mapping and any(f in mapping for f in ['cia1', 'cia2', 'total']):
        actual_matches += 2
        expected_matches += 1
        
    confidence = (actual_matches / expected_matches) if expected_matches > 0 else 0.0
    return mapping, min(1.0, confidence), import_type

# ─────────────────────────────────────────────────────────────────────────────
# STUDENT MATCHING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def match_student(cur, row, class_id=None):
    """
    Smart student matcher using fallback priority criteria.
    Returns (student_id, confidence, matched_by_criteria)
    """
    # 1. Gather identifiers from row
    reg_no = str(row.get('register_number') or '').strip()
    email = str(row.get('email') or '').strip()
    name = str(row.get('student_name') or '').strip()
    dob = str(row.get('dob') or '').strip() # expected format YYYY-MM-DD
    
    if not reg_no and not email and not name:
        return None, 0.0, 'No identifier'
        
    # Helper to verify student matches selected class
    def verify_class(s_id):
        if not class_id:
            return True
        cur.execute("SELECT class_id FROM Students WHERE user_id = %s", (s_id,))
        res = cur.fetchone()
        return res and res['class_id'] == class_id

    # PRIORITY 1: Register Number
    if reg_no:
        cur.execute("SELECT user_id FROM Students WHERE register_number = %s", (reg_no,))
        res = cur.fetchone()
        if res:
            s_id = res['user_id']
            if verify_class(s_id):
                return s_id, 1.0, 'Register Number'
            else:
                return s_id, 0.8, 'Register Number (Class mismatch)'

    # PRIORITY 2: Email
    if email:
        cur.execute("SELECT id FROM Users WHERE email = %s AND role = 'student'", (email,))
        res = cur.fetchone()
        if res:
            s_id = res['id']
            if verify_class(s_id):
                return s_id, 0.95, 'Email'
            else:
                return s_id, 0.75, 'Email (Class mismatch)'

    # PRIORITY 3: Name + DOB
    if name and dob:
        try:
            parsed_dob = datetime.strptime(dob, "%Y-%m-%d").date()
        except:
            parsed_dob = None
            
        if parsed_dob:
            cur.execute("""
                SELECT u.id FROM Users u 
                JOIN Students s ON u.id = s.user_id 
                WHERE LOWER(u.name) = LOWER(%s) AND u.dob = %s
            """, (name, parsed_dob))
            res = cur.fetchone()
            if res:
                s_id = res['id']
                if verify_class(s_id):
                    return s_id, 0.9, 'Name + DOB'
                else:
                    return s_id, 0.7, 'Name + DOB (Class mismatch)'

    # PRIORITY 4: Name + Class
    if name and class_id:
        cur.execute("""
            SELECT u.id FROM Users u 
            JOIN Students s ON u.id = s.user_id 
            WHERE LOWER(u.name) = LOWER(%s) AND s.class_id = %s
        """, (name, class_id))
        res = cur.fetchall()
        if len(res) == 1:
            return res[0]['id'], 0.8, 'Name + Class'
        elif len(res) > 1:
            return None, 0.0, 'Multiple matches for Name + Class'

    # PRIORITY 5: Name Loose Match
    if name:
        cur.execute("""
            SELECT u.id FROM Users u 
            JOIN Students s ON u.id = s.user_id 
            WHERE LOWER(u.name) LIKE LOWER(%s)
        """, (f"%{name}%",))
        res = cur.fetchall()
        if len(res) == 1:
            s_id = res[0]['id']
            if verify_class(s_id):
                return s_id, 0.6, 'Name Loose Match'

    return None, 0.0, 'No match found'

# ─────────────────────────────────────────────────────────────────────────────
# BACKGROUND VALIDATION & RUNNERS
# ─────────────────────────────────────────────────────────────────────────────

def _async_validate(job_id, column_mapping, duplicate_strategy):
    """Background thread function for full row validation."""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Load Job info
        cur.execute("SELECT * FROM ImportJobs WHERE id = %s", (job_id,))
        job = cur.fetchone()
        if not job:
            return
            
        class_id = job['class_id']
        subject_id = job['subject_id']
        import_type = job['import_type']
        
        # Log start
        cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Validation engine started.')", (job_id,))
        
        # Check Semester Locking
        cur.execute("SELECT is_locked FROM SemesterLocks WHERE semester_id = %s", (job['semester_id'],))
        lock_info = cur.fetchone()
        if lock_info and lock_info['is_locked']:
            cur.execute("""
                INSERT INTO ImportErrors (job_id, row_number, error_message, suggestion, error_type)
                VALUES (%s, 0, 'Semester is locked.', 'Ask Administrator to unlock semester.', 'VALIDATION')
            """, (job_id,))
            cur.execute("UPDATE ImportJobs SET status = 'failed' WHERE id = %s", (job_id,))
            return

        imported_data = job['imported_data']
        total_rows = len(imported_data)
        
        valid_count = 0
        warning_count = 0
        error_count = 0
        
        # Build inverse mapping: DB fields keys -> spreadsheet column header names
        inv_map = {}
        for db_f, csv_h in column_mapping.items():
            if csv_h:
                inv_map[db_f] = csv_h
                
        # Cache list of valid students in this class for fast lookup
        cur.execute("SELECT user_id, register_number FROM Students WHERE class_id = %s", (class_id,))
        class_students = {s['register_number']: s['user_id'] for s in cur.fetchall()}
        
        # Clear previous rows/errors for this job in case of retry
        cur.execute("DELETE FROM ImportRows WHERE job_id = %s", (job_id,))
        cur.execute("DELETE FROM ImportErrors WHERE job_id = %s", (job_id,))
        
        # Process row-by-row
        for idx, row in enumerate(imported_data):
            row_num = row['_row_number']
            
            # Map row details using column mappings
            mapped_row = {}
            for db_f, csv_h in inv_map.items():
                mapped_row[db_f] = row.get(csv_h)
                
            start_time = time.time()
            
            row_errors = []
            row_warnings = []
            
            # 1. Student Matcher
            student_id, match_conf, match_by = match_student(cur, mapped_row, class_id)
            if not student_id:
                row_errors.append(f"Student not found in selected Class. Match Strategy: {match_by}.")
            
            # 2. Validation by Import Type
            if import_type == 'attendance' or import_type == 'combined':
                att_status = str(mapped_row.get('attendance_status') or '').strip().upper()
                if att_status:
                    # Clean/normalize status
                    # CHECK: 'P', 'AB', 'SR', 'DL'
                    valid_att_codes = {
                        'P': 'P', 'PRESENT': 'P', '1': 'P', 'TRUE': 'P',
                        'A': 'AB', 'ABSENT': 'AB', '0': 'AB', 'FALSE': 'AB', 'AB': 'AB',
                        'DL': 'DL', 'DUTY LEAVE': 'DL', 'L': 'DL', 'LATE': 'DL',
                        'S': 'SR', 'SICK': 'SR', 'MEDICAL LEAVE': 'SR', 'SR': 'SR'
                    }
                    if att_status not in valid_att_codes:
                        row_errors.append(f"Invalid attendance value '{att_status}'.")
                    else:
                        mapped_row['attendance_status'] = valid_att_codes[att_status]
                else:
                    if import_type == 'attendance':
                        row_errors.append("Missing required Attendance status.")
                        
            if import_type == 'marks' or import_type == 'combined':
                # Check all mapped marks columns
                marks_keys = ['cia1', 'cia2', 'cia3', 'assignment', 'lab', 'project', 'quiz', 'seminar', 'total']
                for key in marks_keys:
                    val = mapped_row.get(key)
                    if val is not None and str(val).strip() != "":
                        try:
                            fval = float(val)
                            if fval < 0:
                                row_errors.append(f"Marks for '{key}' cannot be negative ({fval}).")
                            elif fval > 100:
                                # Standard cap check, can be customized
                                row_warnings.append(f"Marks for '{key}' exceeds 100 ({fval}). Verify if correct.")
                            mapped_row[key] = fval
                        except ValueError:
                            row_errors.append(f"Invalid numeric format for '{key}' marks ('{val}').")
                            
            # 3. Duplicate checks within sheet
            reg_val = mapped_row.get('register_number')
            # Check for multiple instances of this register number in the spreadsheet
            dups = [r for r in imported_data if str(r.get(inv_map.get('register_number'))).strip() == str(reg_val).strip()]
            if len(dups) > 1:
                row_warnings.append("Duplicate student register number in sheet.")

            # Determine Row Status
            status = 'valid'
            if row_errors:
                status = 'error'
                error_count += 1
            elif row_warnings:
                status = 'warning'
                warning_count += 1
            else:
                status = 'valid'
                valid_count += 1
                
            duration = int((time.time() - start_time) * 1000)
            
            # Save ImportRows
            cur.execute("""
                INSERT INTO ImportRows (job_id, row_number, raw_data, student_id, validation_status, error_message, processing_time_ms)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (job_id, row_num, json.dumps(mapped_row), student_id, status, "; ".join(row_errors), duration))
            
            # Save ImportErrors
            for err in row_errors:
                cur.execute("""
                    INSERT INTO ImportErrors (job_id, row_number, student_identifier, error_message, suggestion, error_type)
                    VALUES (%s, %s, %s, %s, %s, 'VALIDATION')
                """, (job_id, row_num, reg_val, err, 'Correct details in the Excel sheet or resolve student mapping.'))
                
            for wrn in row_warnings:
                cur.execute("""
                    INSERT INTO ImportErrors (job_id, row_number, student_identifier, error_message, suggestion, error_type)
                    VALUES (%s, %s, %s, %s, %s, 'VALIDATION')
                """, (job_id, row_num, reg_val, f"[Warning] {wrn}", 'Double check details.'))
                
            # Update Progress percent
            pct = int(((idx + 1) / total_rows) * 100)
            cur.execute("UPDATE ImportJobs SET progress_percent = %s WHERE id = %s", (pct, job_id))
            
        # Update Job summary
        cur.execute("""
            UPDATE ImportJobs 
            SET status = 'validated', valid_rows = %s, warning_rows = %s, error_rows = %s, progress_percent = 100
            WHERE id = %s
        """, (valid_count, warning_count, error_count, job_id))
        
        cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Validation engine completed successfully.')", (job_id,))
        conn.commit()
        
    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Error in background validation: {str(e)}")
        # Log to ImportLogs and fail job
        try:
            cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'ERROR', %s)", (job_id, f"Validation crashed: {str(e)}"))
            cur.execute("UPDATE ImportJobs SET status = 'failed' WHERE id = %s", (job_id,))
            conn.commit()
        except:
            pass
    finally:
        conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# ROUTE DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────

@import_bp.route('/upload', methods=['POST'])
@token_required(allowed_roles=['faculty', 'admin'])
def upload_file(current_user):
    file = request.files.get('file')
    class_id = request.form.get('class_id')
    subject_id = request.form.get('subject_id')
    import_type = request.form.get('import_type', 'auto') # marks, attendance, combined, auto
    academic_year = request.form.get('academic_year')
    exam_type = request.form.get('exam_type') # e.g. internal, lab, assignment
    hour = request.form.get('hour')
    date_val = request.form.get('date')
    
    if not file:
        return jsonify({'message': 'No spreadsheet file provided.'}), 400
    if not class_id or not subject_id:
        return jsonify({'message': 'class_id and subject_id are required fields.'}), 400
        
    file_bytes = file.read()
    file_name = file.filename
    
    try:
        # 1. Parse spreadsheet rows
        headers, records = parse_file(file_bytes, file_name)
        if not headers:
            return jsonify({'message': 'Uploaded file contains no headers or records.'}), 400
            
        # 2. AI Column Detection
        mapping, confidence, detected_type = detect_column_mappings(headers, import_type)
        
        # 3. Create Job
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Resolve class metadata (institution, department, semester)
        cur.execute("""
            SELECT c.department_id, c.semester_id, d.institution_id 
            FROM Classes c
            JOIN Departments d ON c.department_id = d.id
            WHERE c.id = %s
        """, (class_id,))
        class_meta = cur.fetchone()
        
        inst_id = class_meta['institution_id'] if class_meta else None
        dept_id = class_meta['department_id'] if class_meta else None
        sem_id = class_meta['semester_id'] if class_meta else None
        
        rollback_token = str(uuid.uuid4())
        
        # Get Client Metadata
        meta = {
            'ip_address': request.remote_addr,
            'user_agent': request.user_agent.string if request.user_agent else 'Unknown',
            'browser': request.user_agent.browser if request.user_agent else 'Unknown',
            'platform': request.user_agent.platform if request.user_agent else 'Unknown'
        }
        
        cur.execute("""
            INSERT INTO ImportJobs 
            (faculty_id, institution_id, department_id, semester_id, class_id, subject_id, 
             academic_year, exam_type, hour, date, file_name, import_type, status, 
             total_rows, rollback_token, column_mapping, imported_data, metadata)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'uploaded', %s, %s, %s, %s, %s)
            RETURNING id
        """, (current_user['user_id'], inst_id, dept_id, sem_id, class_id, subject_id,
              academic_year, exam_type, int(hour) if hour else None, date_val if date_val else None,
              file_name, detected_type, len(records), rollback_token, json.dumps(mapping),
              json.dumps(records), json.dumps(meta)))
              
        job_id = cur.fetchone()['id']
        
        # Add Audit log
        cur.execute("""
            INSERT INTO ImportHistory (job_id, faculty_id, action, details, ip_address, user_agent)
            VALUES (%s, %s, 'UPLOAD', %s, %s, %s)
        """, (job_id, current_user['user_id'], json.dumps({'file_name': file_name, 'rows': len(records)}),
              meta['ip_address'], meta['user_agent']))
              
        conn.commit()
        conn.close()
        
        # Log activity
        log_activity(current_user['user_id'], 'IMPORT_FILE_UPLOADED', 'import_job', job_id, 
                     new_data={'file_name': file_name, 'rows': len(records)})
                     
        return jsonify({
            'message': 'File uploaded successfully.',
            'job_id': job_id,
            'headers': headers,
            'detected_mapping': mapping,
            'detected_type': detected_type,
            'confidence': round(confidence * 100, 1),
            'preview_rows': records[:5],
            'total_rows': len(records)
        }), 201
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'message': f'Upload processing failed: {str(e)}'}), 500

@import_bp.route('/validate', methods=['POST'])
@token_required(allowed_roles=['faculty', 'admin'])
def start_validation(current_user):
    data = request.json or {}
    job_id = data.get('job_id')
    column_mapping = data.get('column_mapping') # updated mapping from UI
    duplicate_strategy = data.get('duplicate_strategy', 'replace') # replace, skip, merge, update_empty
    
    if not job_id:
        return jsonify({'message': 'job_id is required.'}), 400
        
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ImportJobs WHERE id = %s", (job_id,))
    job = cur.fetchone()
    if not job:
        conn.close()
        return jsonify({'message': 'Job not found.'}), 404
        
    # Verify authorization (only assigned faculty can run validations)
    # Check if faculty is assigned to this subject/class
    cur.execute("""
        SELECT id FROM SubjectAssignments WHERE class_id = %s AND teacher_id = %s AND subject_id = %s
    """, (job['class_id'], current_user['user_id'], job['subject_id']))
    if not cur.fetchone():
        # Check if they are class teacher
        cur.execute("""
            SELECT id FROM ClassAssignments WHERE class_id = %s AND teacher_id = %s 
            AND role IN ('class_teacher', 'vice_class_teacher')
        """, (job['class_id'], current_user['user_id']))
        if not cur.fetchone():
            conn.close()
            return jsonify({'message': 'Unauthorized: Not assigned to this subject or class.'}), 403
            
    # Update mapping if provided
    if column_mapping:
        cur.execute("UPDATE ImportJobs SET column_mapping = %s, status = 'validating', progress_percent = 0 WHERE id = %s", 
                    (json.dumps(column_mapping), job_id))
    else:
        cur.execute("UPDATE ImportJobs SET status = 'validating', progress_percent = 0 WHERE id = %s", (job_id,))
        column_mapping = job['column_mapping']
        
    # Add Log entry
    cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Validation process queued.')", (job_id,))
    
    conn.commit()
    conn.close()
    
    # Spawn background validation thread
    t = threading.Thread(target=_async_validate, args=(job_id, column_mapping, duplicate_strategy))
    t.daemon = True
    t.start()
    
    return jsonify({
        'message': 'Validation engine started in background.',
        'job_id': job_id,
        'status': 'validating'
    }), 200

@import_bp.route('/job/<int:job_id>', methods=['GET'])
@token_required(allowed_roles=['faculty', 'admin'])
def get_job_status(current_user, job_id):
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM ImportJobs WHERE id = %s", (job_id,))
    job = cur.fetchone()
    if not job:
        conn.close()
        return jsonify({'message': 'Job not found.'}), 404
        
    # Fetch logs
    cur.execute("SELECT log_level, message, created_at FROM ImportLogs WHERE job_id = %s ORDER BY id ASC", (job_id,))
    logs = cur.fetchall()
    
    # Fetch error counts
    cur.execute("SELECT COUNT(*) as count FROM ImportErrors WHERE job_id = %s AND error_type = 'VALIDATION'", (job_id,))
    errs = cur.fetchone()['count']
    
    conn.close()
    
    return jsonify({
        'job_id': job_id,
        'status': job['status'],
        'progress_percent': job['progress_percent'],
        'total_rows': job['total_rows'],
        'valid_rows': job['valid_rows'],
        'warning_rows': job['warning_rows'],
        'error_rows': job['error_rows'],
        'errors_count': errs,
        'logs': [{'level': l['log_level'], 'message': l['message'], 'time': l['created_at'].isoformat()} for l in logs]
    }), 200

@import_bp.route('/preview', methods=['GET'])
@token_required(allowed_roles=['faculty', 'admin'])
def get_job_preview(current_user):
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'message': 'job_id is required.'}), 400
        
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT r.row_number, r.raw_data, r.validation_status, r.error_message, s.register_number, u.name as student_name
        FROM ImportRows r
        LEFT JOIN Students s ON r.student_id = s.user_id
        LEFT JOIN Users u ON s.user_id = u.id
        WHERE r.job_id = %s
        ORDER BY r.row_number ASC
    """, (job_id,))
    rows = cur.fetchall()
    
    # Fetch errors
    cur.execute("SELECT row_number, error_message, suggestion, error_type FROM ImportErrors WHERE job_id = %s", (job_id,))
    errors = cur.fetchall()
    
    conn.close()
    
    return jsonify({
        'job_id': job_id,
        'rows': [
            {
                'row_number': r['row_number'],
                'register_number': r['register_number'],
                'student_name': r['student_name'],
                'raw_data': r['raw_data'],
                'status': r['validation_status'],
                'error_message': r['error_message']
            } for r in rows
        ],
        'errors': [
            {
                'row_number': e['row_number'],
                'message': e['error_message'],
                'suggestion': e['suggestion'],
                'type': e['error_type']
            } for e in errors
        ]
    }), 200

# ─────────────────────────────────────────────────────────────────────────────
# IMPORT EXECUTION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _async_execute(job_id, current_user_id, conflict_strategy):
    """Background thread function for writing spreadsheet records to target tables."""
    conn = get_db_connection()
    conn.autocommit = False # transactional block
    cur = conn.cursor()
    
    start_time = time.time()
    
    try:
        # Load Job
        cur.execute("SELECT * FROM ImportJobs WHERE id = %s", (job_id,))
        job = cur.fetchone()
        
        class_id = job['class_id']
        subject_id = job['subject_id']
        import_type = job['import_type']
        
        cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Import execution started.')", (job_id,))
        
        # Retrieve parsed rows
        cur.execute("SELECT * FROM ImportRows WHERE job_id = %s AND validation_status != 'error' AND student_id IS NOT NULL", (job_id,))
        valid_rows = cur.fetchall()
        
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        
        previous_records = []
        
        # Process each valid row
        for idx, row in enumerate(valid_rows):
            student_id = row['student_id']
            raw_data = row['raw_data']
            row_num = row['row_number']
            
            # Get student register number
            cur.execute("SELECT register_number FROM Students WHERE user_id = %s", (student_id,))
            s_reg = cur.fetchone()['register_number']
            
            # --- 1. ATTENDANCE EXECUTION ---
            if import_type == 'attendance' or import_type == 'combined':
                att_status = raw_data.get('attendance_status')
                att_date = job['date']
                att_hour = job['hour']
                
                # If combined, check if date/hour is inside row data or job metadata
                row_date = raw_data.get('date') or att_date
                row_hour = raw_data.get('hour') or att_hour
                
                if row_date and row_hour and att_status:
                    # Check if already exists
                    cur.execute("""
                        SELECT id, status FROM Attendance 
                        WHERE student_id = %s AND class_id = %s AND date = %s AND hour = %s AND subject_id = %s
                    """, (student_id, class_id, row_date, int(row_hour), subject_id))
                    existing = cur.fetchone()
                    
                    if existing:
                        if conflict_strategy == 'skip':
                            skipped_count += 1
                            continue
                        elif conflict_strategy == 'replace' or conflict_strategy == 'merge':
                            # Store previous value for rollback snapshot
                            previous_records.append({
                                'table': 'Attendance',
                                'action': 'update',
                                'id': existing['id'],
                                'previous_values': {'status': existing['status']},
                                'new_values': {'status': att_status}
                            })
                            # Update
                            cur.execute("""
                                UPDATE Attendance SET status = %s, updated_by = %s, updated_at = CURRENT_TIMESTAMP
                                WHERE id = %s
                            """, (att_status, current_user_id, existing['id']))
                            
                            # Log to AttendanceLogs
                            cur.execute("""
                                INSERT INTO AttendanceLogs (attendance_id, previous_status, new_status, changed_by)
                                VALUES (%s, %s, %s, %s)
                            """, (existing['id'], existing['status'], att_status, current_user_id))
                            
                            updated_count += 1
                    else:
                        # Insert
                        cur.execute("""
                            INSERT INTO Attendance (student_id, class_id, date, hour, subject_id, status, updated_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            RETURNING id
                        """, (student_id, class_id, row_date, int(row_hour), subject_id, att_status, current_user_id))
                        new_id = cur.fetchone()['id']
                        
                        previous_records.append({
                            'table': 'Attendance',
                            'action': 'insert',
                            'id': new_id
                        })
                        
                        # Log to AttendanceLogs
                        cur.execute("""
                            INSERT INTO AttendanceLogs (attendance_id, previous_status, new_status, changed_by)
                            VALUES (%s, NULL, %s, %s)
                        """, (new_id, att_status, current_user_id))
                        
                        inserted_count += 1

            # --- 2. MARKS EXECUTION ---
            if import_type == 'marks' or import_type == 'combined':
                marks_keys = ['cia1', 'cia2', 'cia3', 'assignment', 'lab', 'project', 'quiz', 'seminar', 'total']
                for key in marks_keys:
                    score = raw_data.get(key)
                    if score is not None:
                        # Check if exists
                        cur.execute("""
                            SELECT id, marks FROM Marks
                            WHERE student_id = %s AND subject_id = %s AND mark_type = %s AND mark_name = %s
                        """, (student_id, subject_id, job['exam_type'] or 'internal', key))
                        existing = cur.fetchone()
                        
                        if existing:
                            if conflict_strategy == 'skip':
                                skipped_count += 1
                                continue
                            elif conflict_strategy == 'update_empty':
                                if existing['marks'] is not None and float(existing['marks']) > 0:
                                    skipped_count += 1
                                    continue
                                    
                            # Update / Replace
                            previous_records.append({
                                'table': 'Marks',
                                'action': 'update',
                                'id': existing['id'],
                                'previous_values': {'marks': float(existing['marks'])},
                                'new_values': {'marks': float(score)}
                            })
                            cur.execute("""
                                UPDATE Marks SET marks = %s, updated_by = %s, updated_at = CURRENT_TIMESTAMP
                                WHERE id = %s
                            """, (score, current_user_id, existing['id']))
                            updated_count += 1
                        else:
                            # Insert
                            cur.execute("""
                                INSERT INTO Marks (student_id, subject_id, mark_type, mark_name, marks, updated_by)
                                VALUES (%s, %s, %s, %s, %s, %s)
                                RETURNING id
                            """, (student_id, subject_id, job['exam_type'] or 'internal', key, score, current_user_id))
                            new_id = cur.fetchone()['id']
                            
                            previous_records.append({
                                'table': 'Marks',
                                'action': 'insert',
                                'id': new_id
                            })
                            inserted_count += 1
                            
            # Update row execution status
            cur.execute("UPDATE ImportRows SET execution_status = 'imported' WHERE id = %s", (row['id'],))
            
            # Progress percent updates
            pct = int(((idx + 1) / len(valid_rows)) * 100)
            cur.execute("UPDATE ImportJobs SET progress_percent = %s WHERE id = %s", (pct, job_id))

        # --- 3. GENERATE AUTOMATIC ANALYTICS ---
        summary_stats = {
            'inserted': inserted_count,
            'updated': updated_count,
            'skipped': skipped_count
        }
        
        if import_type == 'marks' or import_type == 'combined':
            # Gather all marks for this subject
            cur.execute("SELECT marks FROM Marks WHERE subject_id = %s", (subject_id,))
            all_marks = [float(r['marks']) for r in cur.fetchall() if r['marks'] is not None]
            if all_marks:
                summary_stats['marks'] = {
                    'average': round(sum(all_marks) / len(all_marks), 2),
                    'highest': max(all_marks),
                    'lowest': min(all_marks),
                    'total_records': len(all_marks)
                }
                
        if import_type == 'attendance' or import_type == 'combined':
            # Gather overall attendance rate for this class
            cur.execute("SELECT status FROM Attendance WHERE class_id = %s", (class_id,))
            all_att = [r['status'] for r in cur.fetchall()]
            if all_att:
                p_count = sum(1 for a in all_att if a == 'P' or a == 'DL')
                summary_stats['attendance'] = {
                    'attendance_percentage': round((p_count / len(all_att)) * 100, 2),
                    'total_records': len(all_att)
                }
                
        # Save ImportAnalytics
        cur.execute("INSERT INTO ImportAnalytics (job_id, summary_stats) VALUES (%s, %s)", (job_id, json.dumps(summary_stats)))

        # Create sequential ImportVersion
        # Look up max version
        cur.execute("""
            SELECT COALESCE(MAX(version_number), 0) as max_v FROM ImportVersions
            WHERE class_id = %s AND subject_id = %s AND import_type = %s
        """, (class_id, subject_id, import_type))
        next_version = cur.fetchone()['max_v'] + 1
        
        # Set previous versions inactive
        cur.execute("""
            UPDATE ImportVersions SET active = FALSE 
            WHERE class_id = %s AND subject_id = %s AND import_type = %s
        """, (class_id, subject_id, import_type))
        
        # Insert new active version
        cur.execute("""
            INSERT INTO ImportVersions (class_id, subject_id, import_type, exam_type, date, hour, version_number, job_id, active)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
        """, (class_id, subject_id, import_type, job['exam_type'], job['date'], job['hour'], next_version, job_id))

        # --- 4. APPROVAL WORKFLOW RESOLUTION ---
        # Add to HOD approval queue as pending (can be bypassed by admins)
        cur.execute("SELECT role FROM Users WHERE id = %s", (current_user_id,))
        role = cur.fetchone()['role']
        
        if role == 'admin':
            # Auto-approve
            cur.execute("UPDATE ImportJobs SET is_approved = TRUE, approved_by = %s, approved_at = CURRENT_TIMESTAMP WHERE id = %s", 
                        (current_user_id, job_id))
            cur.execute("UPDATE ImportJobs SET status = 'completed' WHERE id = %s", (job_id,))
            cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Import executed and approved immediately by Administrator.')", (job_id,))
        else:
            # Place in HOD review
            cur.execute("INSERT INTO ApprovalQueue (job_id, status) VALUES (%s, 'pending')", (job_id,))
            cur.execute("UPDATE ImportJobs SET status = 'completed', is_approved = FALSE WHERE id = %s", (job_id,))
            cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Import completed. Placed in HOD Approval Queue.')", (job_id,))
            
            # Trigger notifications for Department HODs
            # In our schema, HOD role is typically 'admin' or we notify academic admins
            cur.execute("SELECT id FROM Users WHERE role = 'admin' AND institution_id = %s", (job['institution_id'],))
            admins = cur.fetchall()
            for admin in admins:
                cur.execute("""
                    INSERT INTO Notifications (user_id, title, message, type)
                    VALUES (%s, 'Approval Queue Pending', 'Bulk data import is pending approval.', 'system')
                    RETURNING id
                """, (admin['id'],))
                notif_id = cur.fetchone()['id']
                cur.execute("INSERT INTO ImportNotifications (job_id, notification_id) VALUES (%s, %s)", (job_id, notif_id))

        duration = int((time.time() - start_time) * 1000)
        
        # Save snapshot previous records to job (for 24-hr rollbacks)
        cur.execute("UPDATE ImportJobs SET previous_records = %s, duration_ms = %s WHERE id = %s", 
                    (json.dumps(previous_records), duration, job_id))
        
        # Add Log entry
        cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Import completed successfully.')", (job_id,))
        
        # Add Audit trail log
        cur.execute("""
            INSERT INTO ImportHistory (job_id, faculty_id, action, details, ip_address, user_agent)
            VALUES (%s, %s, 'EXECUTE', %s, NULL, NULL)
        """, (job_id, current_user_id, json.dumps(summary_stats)))
        
        conn.commit()
        
    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Error in background execution: {str(e)}")
        traceback.print_exc()
        try:
            cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'ERROR', %s)", (job_id, f"Execution failed: {str(e)}"))
            cur.execute("UPDATE ImportJobs SET status = 'failed' WHERE id = %s", (job_id,))
            conn.commit()
        except:
            pass
    finally:
        conn.close()

@import_bp.route('/execute', methods=['POST'])
@token_required(allowed_roles=['faculty', 'admin'])
def execute_import(current_user):
    data = request.json or {}
    job_id = data.get('job_id')
    conflict_strategy = data.get('duplicate_strategy', 'replace') # replace, skip, merge, update_empty
    
    if not job_id:
        return jsonify({'message': 'job_id is required.'}), 400
        
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT status FROM ImportJobs WHERE id = %s", (job_id,))
    job = cur.fetchone()
    if not job:
        conn.close()
        return jsonify({'message': 'Job not found.'}), 404
        
    if job['status'] not in ['validated', 'failed']:
        conn.close()
        return jsonify({'message': f'Job cannot be executed. Status is: {job["status"]}'}), 400
        
    cur.execute("UPDATE ImportJobs SET status = 'importing', progress_percent = 0 WHERE id = %s", (job_id,))
    conn.commit()
    conn.close()
    
    # Spawn background thread for execution
    t = threading.Thread(target=_async_execute, args=(job_id, current_user['user_id'], conflict_strategy))
    t.daemon = True
    t.start()
    
    return jsonify({
        'message': 'Data insertion started in background.',
        'job_id': job_id,
        'status': 'importing'
    }), 200

# ─────────────────────────────────────────────────────────────────────────────
# ROLLBACK ENGINE & SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

@import_bp.route('/rollback', methods=['POST'])
@token_required(allowed_roles=['faculty', 'admin'])
def rollback_import(current_user):
    data = request.json or {}
    rollback_token = data.get('rollback_token')
    job_id = data.get('job_id')
    
    if not rollback_token and not job_id:
        return jsonify({'message': 'rollback_token or job_id is required.'}), 400
        
    conn = get_db_connection()
    conn.autocommit = False # Transaction control
    cur = conn.cursor()
    
    try:
        # Load Job
        if rollback_token:
            cur.execute("SELECT * FROM ImportJobs WHERE rollback_token = %s", (rollback_token,))
        else:
            cur.execute("SELECT * FROM ImportJobs WHERE id = %s", (job_id,))
            
        job = cur.fetchone()
        if not job:
            conn.close()
            return jsonify({'message': 'Import job not found.'}), 404
            
        if job['status'] != 'completed':
            conn.close()
            return jsonify({'message': f'Job cannot be rolled back. Status is {job["status"]}.'}), 400
            
        # Verify 24 hour limits (unless administrator)
        created_time = job['created_at']
        diff = datetime.now() - created_time
        if diff.total_seconds() > 86400 and current_user['role'] != 'admin':
            conn.close()
            return jsonify({'message': 'Rollback time window (24 hours) has expired. Contact Admin.'}), 400
            
        previous_records = job['previous_records']
        if not previous_records:
            conn.close()
            return jsonify({'message': 'No snapshots or previous records saved for this job. Cannot rollback.'}), 400
            
        # Perform rollback: reverse snapshots
        # Iterate in reverse order to ensure constraint safety
        for rec in reversed(previous_records):
            tbl = rec['table']
            act = rec['action']
            r_id = rec['id']
            
            if act == 'insert':
                # Delete newly inserted records
                if tbl == 'Attendance':
                    # Delete logs first
                    cur.execute("DELETE FROM AttendanceLogs WHERE attendance_id = %s", (r_id,))
                    cur.execute("DELETE FROM Attendance WHERE id = %s", (r_id,))
                elif tbl == 'Marks':
                    cur.execute("DELETE FROM Marks WHERE id = %s", (r_id,))
            elif act == 'update':
                # Restore previous values
                prev_vals = rec['previous_values']
                if tbl == 'Attendance':
                    cur.execute("UPDATE Attendance SET status = %s WHERE id = %s", (prev_vals['status'], r_id))
                    # Log change to logs
                    cur.execute("""
                        INSERT INTO AttendanceLogs (attendance_id, previous_status, new_status, changed_by)
                        VALUES (%s, %s, %s, %s)
                    """, (r_id, rec['new_values']['status'], prev_vals['status'], current_user['user_id']))
                elif tbl == 'Marks':
                    cur.execute("UPDATE Marks SET marks = %s WHERE id = %s", (prev_vals['marks'], r_id))

        # Delete notifications sent
        cur.execute("SELECT notification_id FROM ImportNotifications WHERE job_id = %s", (job['id'],))
        notifs = cur.fetchall()
        for notif in notifs:
            cur.execute("DELETE FROM Notifications WHERE id = %s", (notif['notification_id'],))
        cur.execute("DELETE FROM ImportNotifications WHERE job_id = %s", (job['id'],))

        # Delete Version
        cur.execute("DELETE FROM ImportVersions WHERE job_id = %s", (job['id'],))
        
        # Restore prior active version
        cur.execute("""
            SELECT id FROM ImportVersions 
            WHERE class_id = %s AND subject_id = %s AND import_type = %s 
            ORDER BY version_number DESC LIMIT 1
        """, (job['class_id'], job['subject_id'], job['import_type']))
        prior_v = cur.fetchone()
        if prior_v:
            cur.execute("UPDATE ImportVersions SET active = TRUE WHERE id = %s", (prior_v['id'],))

        # Update Job Status
        cur.execute("UPDATE ImportJobs SET status = 'rolled_back' WHERE id = %s", (job['id'],))
        
        # Log to ImportLogs
        cur.execute("INSERT INTO ImportLogs (job_id, log_level, message) VALUES (%s, 'INFO', 'Job import rolled back successfully.')", (job['id'],))
        
        # Log to ImportHistory
        cur.execute("""
            INSERT INTO ImportHistory (job_id, faculty_id, action, details, ip_address, user_agent)
            VALUES (%s, %s, 'ROLLBACK', %s, NULL, NULL)
        """, (job['id'], current_user['user_id'], json.dumps({'timestamp': datetime.now().isoformat()})))

        conn.commit()
        conn.close()
        
        log_activity(current_user['user_id'], 'IMPORT_ROLLED_BACK', 'import_job', job['id'])
        
        return jsonify({
            'message': 'Rollback completed successfully. All modified and inserted records restored to previous state.',
            'job_id': job['id']
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Error during rollback: {str(e)}")
        traceback.print_exc()
        return jsonify({'message': f'Rollback failed: {str(e)}'}), 500

# ─────────────────────────────────────────────────────────────────────────────
# HISTORY & ERROR REPORTS
# ─────────────────────────────────────────────────────────────────────────────

@import_bp.route('/history', methods=['GET'])
@token_required(allowed_roles=['faculty', 'admin'])
def get_import_history(current_user):
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Filter by user if faculty, admins see all
    if current_user['role'] == 'admin':
        cur.execute("""
            SELECT j.id, j.file_name, j.import_type, j.status, j.total_rows, j.valid_rows, j.error_rows, 
                   j.created_at, u.name as faculty_name, c.section, sub.name as subject_name
            FROM ImportJobs j
            JOIN Users u ON j.faculty_id = u.id
            JOIN Classes c ON j.class_id = c.id
            JOIN Subjects sub ON j.subject_id = sub.id
            ORDER BY j.created_at DESC
        """)
    else:
        cur.execute("""
            SELECT j.id, j.file_name, j.import_type, j.status, j.total_rows, j.valid_rows, j.error_rows, 
                   j.created_at, c.section, sub.name as subject_name
            FROM ImportJobs j
            JOIN Classes c ON j.class_id = c.id
            JOIN Subjects sub ON j.subject_id = sub.id
            WHERE j.faculty_id = %s
            ORDER BY j.created_at DESC
        """, (current_user['user_id'],))
        
    jobs = cur.fetchall()
    conn.close()
    
    return jsonify([
        {
            'job_id': j['id'],
            'file_name': j['file_name'],
            'import_type': j['import_type'],
            'status': j['status'],
            'total_rows': j['total_rows'],
            'valid_rows': j['valid_rows'],
            'error_rows': j['error_rows'],
            'faculty_name': j.get('faculty_name') or current_user['name'],
            'class_section': j['section'],
            'subject_name': j['subject_name'],
            'created_at': j['created_at'].isoformat()
        } for j in jobs
    ]), 200

# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE DYNAMIC GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

@import_bp.route('/download-template', methods=['GET'])
@token_required(allowed_roles=['faculty', 'admin'])
def download_template(current_user):
    class_id = request.args.get('class_id')
    subject_id = request.args.get('subject_id')
    template_type = request.args.get('type', 'marks') # marks, attendance, combined
    exam_type = request.args.get('exam_type')
    
    if not class_id or not subject_id:
        return jsonify({'message': 'class_id and subject_id are required fields.'}), 400
        
    conn = get_db_connection()
    cur = conn.cursor()
    
    # 1. Fetch Students
    cur.execute("""
        SELECT s.register_number, u.name 
        FROM Students s
        JOIN Users u ON s.user_id = u.id
        WHERE s.class_id = %s
        ORDER BY s.register_number ASC
    """, (class_id,))
    students = cur.fetchall()
    
    conn.close()
    
    # Create dynamic openpyxl workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Academic Data Template"
    
    # Styles
    hdr_font = openpyxl.styles.Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    hdr_fill = openpyxl.styles.PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid') # premium violet color
    align_center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # Determine columns based on type
    headers = ["Register_No", "Student_Name"]
    if template_type == 'attendance':
        headers += ["Attendance_Status", "Date", "Hour"]
    elif template_type == 'combined':
        headers += ["CIA1", "CIA2", "Attendance_Status"]
    else: # marks
        headers += ["CIA1", "CIA2", "CIA3", "Assignment", "Lab", "Project", "Quiz", "Seminar"]
        
    # Write headers
    for idx, h in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = align_center
        
    # Write dynamic student records
    for row_idx, stud in enumerate(students, 2):
        sheet.cell(row=row_idx, column=1, value=stud['register_number'])
        sheet.cell(row=row_idx, column=2, value=stud['name'])
        
    # Auto-adjust column widths
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max(max_len + 3, 15)
        
    # Save to dynamic buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"EduNexus_{template_type}_template_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ─────────────────────────────────────────────────────────────────────────────
# OFFLINE EXPORT ENGINE
# ─────────────────────────────────────────────────────────────────────────────

@import_bp.route('/export', methods=['GET'])
@token_required(allowed_roles=['faculty', 'admin'])
def export_academic_data(current_user):
    class_id = request.args.get('class_id')
    subject_id = request.args.get('subject_id')
    export_type = request.args.get('type', 'marks')
    
    if not class_id or not subject_id:
        return jsonify({'message': 'class_id and subject_id are required fields.'}), 400
        
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get class details
    cur.execute("""
        SELECT c.section, d.name as department 
        FROM Classes c
        JOIN Departments d ON c.department_id = d.id
        WHERE c.id = %s
    """, (class_id,))
    class_det = cur.fetchone()
    
    # Get subject code
    cur.execute("SELECT name, code FROM Subjects WHERE id = %s", (subject_id,))
    sub_det = cur.fetchone()
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Academic Ledger"
    
    # Styles
    hdr_font = openpyxl.styles.Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    hdr_fill = openpyxl.styles.PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Sleek dark
    align_center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    if export_type == 'attendance':
        # Retrieve all attendance records
        cur.execute("""
            SELECT s.register_number, u.name, a.date, a.hour, a.status
            FROM Attendance a
            JOIN Students s ON a.student_id = s.user_id
            JOIN Users u ON s.user_id = u.id
            WHERE a.class_id = %s AND a.subject_id = %s
            ORDER BY a.date DESC, a.hour ASC, s.register_number ASC
        """, (class_id, subject_id))
        records = cur.fetchall()
        
        headers = ["Register_No", "Student_Name", "Date", "Hour", "Attendance_Status"]
        for idx, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = align_center
            
        for row_idx, r in enumerate(records, 2):
            sheet.cell(row=row_idx, column=1, value=r['register_number'])
            sheet.cell(row=row_idx, column=2, value=r['name'])
            sheet.cell(row=row_idx, column=3, value=r['date'].strftime('%Y-%m-%d') if r['date'] else '')
            sheet.cell(row=row_idx, column=4, value=r['hour'])
            sheet.cell(row=row_idx, column=5, value=r['status'])
    else:
        # Retrieve all marks
        cur.execute("""
            SELECT s.register_number, u.name, m.mark_name, m.marks
            FROM Students s
            JOIN Users u ON s.user_id = u.id
            LEFT JOIN Marks m ON s.user_id = m.student_id AND m.subject_id = %s
            WHERE s.class_id = %s
            ORDER BY s.register_number ASC
        """, (subject_id, class_id))
        records = cur.fetchall()
        
        # Pivot manually
        # student -> mark_name -> marks
        student_map = {}
        mark_names = set()
        for r in records:
            reg = r['register_number']
            if reg not in student_map:
                student_map[reg] = {
                    'register_number': reg,
                    'name': r['name'],
                    'marks': {}
                }
            if r['mark_name']:
                student_map[reg]['marks'][r['mark_name']] = float(r['marks'])
                mark_names.add(r['mark_name'])
                
        sorted_marks = sorted(list(mark_names))
        headers = ["Register_No", "Student_Name"] + sorted_marks
        
        for idx, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = align_center
            
        for row_idx, (reg, s_data) in enumerate(student_map.items(), 2):
            sheet.cell(row=row_idx, column=1, value=s_data['register_number'])
            sheet.cell(row=row_idx, column=2, value=s_data['name'])
            for col_idx, m_name in enumerate(sorted_marks, 3):
                sheet.cell(row=row_idx, column=col_idx, value=s_data['marks'].get(m_name, ""))

    # Auto-adjust column widths
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max(max_len + 3, 15)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    conn.close()
    
    filename = f"EduNexus_Export_{sub_det['code']}_{class_det['section']}_{export_type}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
